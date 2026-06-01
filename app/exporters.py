"""Exportadores genéricos (CSV / XLSX / PDF) a partir de headers + linhas.

Recebe exatamente o que está na tela (cabeçalhos e células já formatadas) e
gera o arquivo. Reutilizável por qualquer aba/tabela do painel.
"""
import csv
import html
import io
from datetime import datetime


def to_csv(headers: list[str], rows: list[list]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")  # ; é amigável p/ Excel pt-BR
    w.writerow(headers)
    for r in rows:
        w.writerow(["" if c is None else c for c in r])
    return ("﻿" + buf.getvalue()).encode("utf-8")  # BOM p/ acento no Excel


def to_xlsx(headers: list[str], rows: list[list], title: str = "Dados") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Dados")[:31]
    ws.append(headers)
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="C8961E")
    for c in ws[1]:
        c.font = bold; c.fill = fill; c.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append(["" if c is None else c for c in r])
    # largura automática simples
    for i, h in enumerate(headers, 1):
        maxlen = max([len(str(h))] + [len(str(r[i-1])) for r in rows if i-1 < len(r)] or [0])
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max(maxlen + 2, 10), 60)
    ws.freeze_panes = "A2"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def to_pdf(headers: list[str], rows: list[list], title: str = "Extrato") -> bytes:
    from weasyprint import HTML

    ths = "".join(f"<th>{html.escape(str(h))}</th>" for h in headers)
    trs = ""
    for r in rows:
        tds = "".join(f"<td>{html.escape('' if c is None else str(c))}</td>" for c in r)
        trs += f"<tr>{tds}</tr>"
    hoje = datetime.now().strftime("%d/%m/%Y %H:%M")
    doc = f"""<!doctype html><html><head><meta charset="utf-8"><style>
      @page {{ size: A4 landscape; margin: 1.2cm; @bottom-right {{ content: "Página " counter(page) " de " counter(pages); font-size: 8px; color:#888; }} }}
      body {{ font-family: Arial, sans-serif; font-size: 9px; color:#222; }}
      h1 {{ font-size: 14px; margin:0 0 2px; }}
      .meta {{ color:#666; font-size:9px; margin-bottom:10px; }}
      table {{ width:100%; border-collapse: collapse; }}
      th, td {{ border:1px solid #ddd; padding:3px 5px; text-align:left; }}
      th {{ background:#C8961E; color:#fff; }}
      tr:nth-child(even) td {{ background:#fafafa; }}
    </style></head><body>
      <h1>{html.escape(title)}</h1>
      <div class="meta">Gerado em {hoje} · {len(rows)} registro(s)</div>
      <table><thead><tr>{ths}</tr></thead><tbody>{trs}</tbody></table>
    </body></html>"""
    return HTML(string=doc).write_pdf()


def build(fmt: str, headers: list[str], rows: list[list], title: str = "Extrato") -> tuple[bytes, str, str]:
    """Retorna (conteudo, media_type, extensao)."""
    fmt = (fmt or "").lower()
    if fmt == "csv":
        return to_csv(headers, rows), "text/csv; charset=utf-8", "csv"
    if fmt == "xlsx":
        return (to_xlsx(headers, rows, title),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx")
    if fmt == "pdf":
        return to_pdf(headers, rows, title), "application/pdf", "pdf"
    raise ValueError(f"Formato inválido: {fmt}")
